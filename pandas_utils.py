import pandas as pd
from openpyxl import load_workbook
from typing import Union
import os
import warnings
from utils import num_cols
import numpy as np


def read_given_years_to_df(years: Union[str, list]) -> pd.DataFrame:
    """
    Function to load data for given years.
    """

    # Handling list or str input
    if isinstance(years, str):
        years = [years]

    datasets = {}
    for year in years:
        this_year_filename = find_data_for_year(year)
        datasets[year] = read_xlsx_to_df(this_year_filename)

    # Handling 2020 column naming differences (one column name changed vs 2018 and 2019)
    datasets['2020'] = datasets['2020'].rename(
        columns={'Annual Time spent at sea [hours]': 'Annual Total time spent at sea [hours]',
                 'Time spent at sea [hours]': 'Total time spent at sea [hours]'})

    # Merging the dfs
    merged_df = pd.concat([df for df in datasets.values()], ignore_index=True)

    return merged_df


def find_data_for_year(year: str) -> str:
    """
    Finds datasets for a given year.
    Assumes filename starts with year and ends with .xlsx
    """
    files_found = []
    data_dir = os.getcwd() + '/data'
    for file in os.listdir(data_dir):
        if file.startswith(year) and file.endswith('.xlsx'):
            files_found.append('data/'+file)

    num_files_found = len(files_found)

    if num_files_found == 1:  # should have a single .xlsx per year
        return files_found[0]
    elif num_files_found == 0:
        raise ValueError(f"No .xlsx files found for year {year}")
    else:
        raise ValueError(f"Too many .xlsx files ({num_files_found}) found for year {year}. "
                         f"One file per year required.")


def read_xlsx_to_df(filename: str) -> pd.DataFrame:
    """
    Reads xlsx file to pandas dataframe.
    Note: Assumes data is in the first/only sheet
    """

    # TODO: this loading is too slow - look into a bytes method to do this
    # TODO: add in some data munging/cleaning where possible
    # Loading a workbook whilst suppressing openpyxl warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(filename)

    df = pd.DataFrame(wb[wb.sheetnames[0]].values)

    # Set df columns as third header row
    df.columns = df.iloc[2].values

    # First two header rows not needed
    df = df.iloc[3:]

    return df


def clean_dataset(df: pd.DataFrame) -> pd.DataFrame:
    # Converting numeric columns - the errors='coerce' option replaces non-numeric strings with nans
    number_cols = num_cols()
    df[number_cols] = df[number_cols].apply(pd.to_numeric, errors='coerce', axis=1)

    # Replacing any error values
    # TODO: errors='coerce' should take care of this automatically anyway - double check and delete below if so
    for bad_value in ['Not Applicable', 'N/A', 'NA', 'None', 'Division by zero!']:
        df = df.replace(bad_value, np.NaN)

    # Cleaning the date columns
    for col in ['DoC issue date', 'DoC expiry date']:
        df[col] = df[col].replace('DoC not issued', pd.NaT)
        df[col] = pd.to_datetime(df[col], format='%d/%m/%Y')

    return df
